"""
Automação de validação de usuários via PyAutoGUI.

Fluxo geral:
    1. Abre a planilha Excel com os cadastros (pasta cadastrohes).
    2. Lê todos os usuários (até encontrar uma linha vazia).
    3. Abre o Google Chrome no site de login.
    4. Para cada usuário: realiza login, confirma, faz logout.
    5. Gera um relatório final em Downloads (resultado_validacao.xlsx).

SEGURANÇA (PyAutoGUI FailSafe):
    Mova o mouse para o CANTO SUPERIOR ESQUERDO da tela a qualquer momento
    para INTERROMPER IMEDIATAMENTE a automação (lança FailSafeException).

Requisitos: Python 3.12+, pyautogui, openpyxl, pyperclip, pygetwindow.
"""

import os
import time
import glob
import subprocess

import pyautogui
import pyperclip
import openpyxl

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO DE SEGURANÇA DO PYAUTOGUI
# ---------------------------------------------------------------------------
# FailSafe: mover o mouse para o canto superior esquerdo (0, 0) aborta tudo.
pyautogui.FAILSAFE = True
# Pausa automática aplicada após CADA ação do PyAutoGUI.
pyautogui.PAUSE = 0.5

# ===========================================================================
# BLOCO DE CONFIGURAÇÃO — COORDENADAS DOS ELEMENTOS NA TELA
# ---------------------------------------------------------------------------
# Todas as posições (x, y) ficam centralizadas aqui. Caso o layout do site
# mude, basta ajustar estes valores. NENHUMA coordenada deve ficar espalhada
# pelo restante do código.
#
# Dica para descobrir as coordenadas: rode `python -c "import pyautogui,
# time; time.sleep(3); print(pyautogui.position())"` posicionando o mouse
# sobre o elemento desejado durante os 3 segundos.
# ===========================================================================

# Campo de Login (usuário)
LOGIN_X = 1929
LOGIN_Y = 418

# Campo de Senha
SENHA_X = 1928
SENHA_Y = 508

# Botão "Entrar" (não utilizado: o login é confirmado pressionando ENTER após
# preencher a senha — ver realizar_login). Mantido aqui caso queira voltar a
# clicar num botão no futuro.
BOTAO_ENTRAR_X = 760
BOTAO_ENTRAR_Y = 510

# Menu do usuário (não utilizado: o logout é feito com um único clique direto
# no botão "Sair" — ver realizar_logout). Mantido para uso futuro.
MENU_USUARIO_X = 1500
MENU_USUARIO_Y = 90

# Botão "Sair" / Logout (clique direto que retorna à tela de login)
BOTAO_SAIR_X = 254
BOTAO_SAIR_Y = 360

# ===========================================================================
# BLOCO DE CONFIGURAÇÃO — TEMPOS (em segundos)
# ---------------------------------------------------------------------------
# Ajuste estes valores para deixar a automação mais rápida ou mais lenta.
# ===========================================================================

TEMPO_ENTRE_CLIQUES = 0.8      # Pausa curta entre cliques/digitações
TEMPO_CARREGAMENTO_SITE = 8.0  # Espera o site carregar completamente
TEMPO_APOS_LOGIN = 3.0         # Espera após clicar em Entrar (confirma login)
TEMPO_APOS_LOGOUT = 3.0        # Espera o retorno à tela inicial de login

# Limite de usuários a processar (0 = todos). Útil para um TESTE controlado:
# defina 1 ou 2 para validar coordenadas/tempos sem percorrer a planilha toda.
# Também pode ser passado na linha de comando:  python main.py 2
LIMITE_USUARIOS = 0

# ===========================================================================
# BLOCO DE CONFIGURAÇÃO — CAMINHOS E URLS
# ===========================================================================

# Valores PADRÃO genéricos (estes ficam no repositório público, sem dados reais).
# A configuração real fica em "config_local.py" — um arquivo IGNORADO pelo Git
# (ver .gitignore), que NÃO é enviado ao GitHub. Assim os dados de caminho/site
# permanecem apenas na sua máquina. Veja "config_local.example.py" como modelo.
PASTA_PLANILHA = r"C:\caminho\para\cadastrohes"
PASTA_RELATORIO = r"C:\caminho\para\relatorios"
NOME_RELATORIO = "resultado_validacao.xlsx"
URL_SITE = "https://seu-sistema-de-login.exemplo.com"

# Sobrescreve os valores acima com os dados reais do arquivo local, se existir.
# Ordem de precedência: config_local.py  >  variável de ambiente  >  padrão.
try:
    import config_local  # arquivo local não versionado
    PASTA_PLANILHA = getattr(config_local, "PASTA_PLANILHA", PASTA_PLANILHA)
    PASTA_RELATORIO = getattr(config_local, "PASTA_RELATORIO", PASTA_RELATORIO)
    URL_SITE = getattr(config_local, "URL_SITE", URL_SITE)
except ImportError:
    # Sem config_local.py: tenta variáveis de ambiente (também não versionadas).
    PASTA_PLANILHA = os.environ.get("PLANILHA_DIR", PASTA_PLANILHA)
    PASTA_RELATORIO = os.environ.get("RELATORIO_DIR", PASTA_RELATORIO)
    URL_SITE = os.environ.get("SITE_URL", URL_SITE)

# Caminho do executável do Chrome (tentativas mais comuns no Windows).
CAMINHOS_CHROME = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
]

# ===========================================================================
# BLOCO DE CONFIGURAÇÃO — TESSERACT OCR
# ---------------------------------------------------------------------------
# O OCR é usado em detectar_falha_login() para ler mensagens de erro na tela
# (ex.: "Senha incorreta"). Caminhos comuns do executável no Windows.
# ===========================================================================

CAMINHOS_TESSERACT = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
]

# Pasta local com os dados de idioma (.traineddata). Mantida no projeto para
# não depender de permissão de administrador na pasta do Tesseract.
TESSDATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tessdata")

# Idioma do OCR (precisa do arquivo <idioma>.traineddata em TESSDATA_DIR).
OCR_IDIOMA = "por"

# --- Classificação do resultado do login (por OCR da tela) ---------------------
#
# A detecção agora é POSITIVA: o login só é considerado bem-sucedido se a tela
# mostrar algum sinal de que o usuário ESTÁ logado (TEXTOS_LOGIN_OK). Caso
# contrário, é tratado como falha — assim um login que deu errado não é mais
# marcado como sucesso por engano.
#
# TEXTOS_LOGIN_OK: palavras que aparecem SÓ depois de logar (ex.: o botão
# "Sair", o nome do usuário, "Bem-vindo"). Ajuste conforme o site real.
TEXTOS_LOGIN_OK = ["sair", "bem-vindo", "bem vindo", "logout"]

# Textos de erro específicos — usados apenas para detalhar o motivo da falha.
TEXTOS_SENHA_INCORRETA = ["senha incorreta", "senha inválida", "senha invalida"]
TEXTOS_USUARIO_INVALIDO = ["usuário inválido", "usuario invalido",
                           "usuário não encontrado", "usuario nao encontrado",
                           "usuário ou senha", "usuario ou senha"]

# Modo de depuração do OCR: se True, salva um print da tela e o texto lido a
# cada tentativa de login na pasta "debug_ocr/". Útil para descobrir os textos
# reais do site e calibrar TEXTOS_LOGIN_OK / TEXTOS_* acima.
MODO_DEBUG_OCR = True
PASTA_DEBUG_OCR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "debug_ocr")


# ===========================================================================
# FUNÇÕES AUXILIARES REUTILIZÁVEIS
# ===========================================================================

def clicar(x, y, pausa=TEMPO_ENTRE_CLIQUES):
    """Clica em uma coordenada da tela e aguarda uma pausa curta."""
    pyautogui.click(x, y)
    time.sleep(pausa)


def digitar_texto(texto, pausa=TEMPO_ENTRE_CLIQUES):
    """
    Digita um texto usando a área de transferência (pyperclip + Ctrl+V).

    Esse método é mais confiável que pyautogui.write() para textos longos
    ou com caracteres especiais/acentos, evitando problemas de layout de
    teclado.
    """
    pyperclip.copy(texto)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(pausa)


def limpar_campo():
    """Seleciona tudo (Ctrl+A) e apaga o conteúdo de um campo de texto."""
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")


# ===========================================================================
# ETAPA 1 — PLANILHA DE CADASTROS
# ===========================================================================

def localizar_arquivo_excel():
    """
    Localiza o único arquivo Excel dentro da pasta de cadastros.

    Retorna o caminho completo do arquivo .xlsx (ou .xlsm) encontrado.
    Lança FileNotFoundError caso não exista nenhum.
    """
    padroes = [os.path.join(PASTA_PLANILHA, "*.xlsx"),
               os.path.join(PASTA_PLANILHA, "*.xlsm")]
    arquivos = []
    for padrao in padroes:
        arquivos.extend(glob.glob(padrao))

    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum arquivo Excel encontrado em: {PASTA_PLANILHA}"
        )
    # Retorna o primeiro (espera-se um único arquivo na pasta).
    return arquivos[0]


def abrir_planilha():
    """
    ETAPA 1 — Abre o Microsoft Excel exibindo a planilha de cadastros.

    A abertura é feita pelo comando padrão do sistema (associação do .xlsx
    ao Excel). Retorna o caminho do arquivo para leitura posterior.
    """
    caminho = localizar_arquivo_excel()
    print(f"[INFO] Abrindo planilha no Excel: {caminho}")
    # os.startfile abre o arquivo no programa padrão (Excel).
    os.startfile(caminho)
    time.sleep(5)  # tempo para o Excel abrir visualmente
    return caminho


def ler_usuarios(caminho_planilha):
    """
    ETAPA 1 (continuação) — Lê todas as linhas com dados da planilha.

    Espera as colunas: Número, Nome, Login, Senha (nesta ordem).
    Percorre as linhas a partir da 2ª (a 1ª é o cabeçalho) e PARA assim que
    encontra uma linha completamente vazia. Não usa quantidade fixa de linhas.

    Retorna uma lista de dicionários:
        [{"numero":..., "nome":..., "login":..., "senha":...}, ...]
    """
    # data_only=True para ler valores calculados em vez de fórmulas.
    workbook = openpyxl.load_workbook(caminho_planilha, data_only=True)
    sheet = workbook.active

    usuarios = []
    # min_row=2 pula o cabeçalho. iter_rows percorre até o fim dos dados.
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        # Considera vazia a linha em que TODAS as células são None/"".
        if linha is None or all(
            celula is None or str(celula).strip() == "" for celula in linha
        ):
            break  # linha completamente vazia → encerra a leitura

        numero = linha[0] if len(linha) > 0 else ""
        nome = linha[1] if len(linha) > 1 else ""
        login = linha[2] if len(linha) > 2 else ""
        senha = linha[3] if len(linha) > 3 else ""

        usuarios.append({
            "numero": "" if numero is None else str(numero).strip(),
            "nome": "" if nome is None else str(nome).strip(),
            "login": "" if login is None else str(login).strip(),
            "senha": "" if senha is None else str(senha).strip(),
        })

    workbook.close()
    print(f"[INFO] {len(usuarios)} usuário(s) lido(s) da planilha.")
    return usuarios


# ===========================================================================
# ETAPA 2 — NAVEGADOR / SITE
# ===========================================================================

def encontrar_chrome():
    """Retorna o caminho do executável do Chrome, ou None se não achar."""
    for caminho in CAMINHOS_CHROME:
        if os.path.isfile(caminho):
            return caminho
    return None


def abrir_chrome():
    """
    ETAPA 2 — Abre o Google Chrome já acessando a URL do sistema.

    Usa subprocess para iniciar o Chrome diretamente no endereço desejado.
    Caso não encontre o executável, recorre ao navegador padrão do sistema.
    """
    chrome = encontrar_chrome()
    if chrome:
        print("[INFO] Abrindo Google Chrome...")
        # --start-maximized garante coordenadas consistentes na tela.
        subprocess.Popen([chrome, "--start-maximized", URL_SITE])
    else:
        print("[AVISO] Chrome não encontrado. Usando navegador padrão.")
        os.startfile(URL_SITE)


def acessar_site():
    """
    ETAPA 2 (continuação) — Aguarda o carregamento completo da página.

    Como não há acesso ao DOM (automação por imagem/coordenadas), o
    carregamento é garantido por um tempo de espera configurável.
    """
    print(f"[INFO] Acessando {URL_SITE} — aguardando carregamento...")
    time.sleep(TEMPO_CARREGAMENTO_SITE)


# ===========================================================================
# ETAPA 3 — LOGIN / LOGOUT
# ===========================================================================

# Cache da configuração do OCR (None = ainda não configurado).
_ocr_pronto = None


def configurar_ocr():
    """
    Configura o Tesseract OCR uma única vez (caminho do executável).

    Localiza o tesseract.exe em CAMINHOS_TESSERACT e o registra no pytesseract.
    Retorna True se o OCR está pronto para uso, False caso contrário.
    O resultado fica em cache para não repetir a busca a cada login.
    """
    global _ocr_pronto
    if _ocr_pronto is not None:
        return _ocr_pronto

    try:
        import pytesseract

        # Procura o executável do Tesseract nos caminhos conhecidos.
        executavel = next((p for p in CAMINHOS_TESSERACT if os.path.isfile(p)),
                          None)
        if executavel:
            pytesseract.pytesseract.tesseract_cmd = executavel

        # Aponta o Tesseract para a pasta local de idiomas. Usamos a variável
        # de ambiente TESSDATA_PREFIX (mais confiável no Windows que a opção
        # --tessdata-dir, que sofre com aspas/barras ao ser dividida).
        if os.path.isdir(TESSDATA_DIR):
            os.environ["TESSDATA_PREFIX"] = TESSDATA_DIR

        # Testa rapidamente se o Tesseract responde.
        pytesseract.get_tesseract_version()
        _ocr_pronto = True
        print(f"[INFO] OCR (Tesseract) configurado: "
              f"{executavel or 'no PATH'} | idioma={OCR_IDIOMA}")
    except Exception as erro:
        # Sem Tesseract/pytesseract: a automação segue sem detecção de erro.
        _ocr_pronto = False
        print(f"[AVISO] OCR indisponível ({erro}). Logins serão tratados como "
              f"sucesso quando não houver exceção.")

    return _ocr_pronto


def ler_tela(rotulo_debug=""):
    """
    Captura a tela e retorna o texto lido por OCR (em minúsculas).

    Se MODO_DEBUG_OCR estiver ativo, salva também o print (.png) e o texto
    (.txt) na pasta debug_ocr/, para ajudar a calibrar os textos de detecção.

    Retorna o texto lido, ou None se o OCR não estiver disponível/falhar.
    """
    if not configurar_ocr():
        return None

    try:
        import pytesseract

        # Captura a tela inteira e roda o OCR no idioma configurado. A pasta
        # de idiomas já foi definida via TESSDATA_PREFIX em configurar_ocr().
        screenshot = pyautogui.screenshot()
        texto = pytesseract.image_to_string(screenshot, lang=OCR_IDIOMA).lower()

        # Depuração: salva print + texto lido para inspeção posterior.
        if MODO_DEBUG_OCR:
            try:
                os.makedirs(PASTA_DEBUG_OCR, exist_ok=True)
                # Mantém só caracteres seguros para nome de arquivo.
                nome = "".join(c if c.isalnum() or c in "-_" else "_"
                               for c in (rotulo_debug or "tela"))
                base = os.path.join(PASTA_DEBUG_OCR, nome)
                screenshot.save(base + ".png")
                with open(base + ".txt", "w", encoding="utf-8") as arq:
                    arq.write(texto)
                print(f"    [debug] OCR salvo em: {base}.png / .txt")
            except Exception as erro_dbg:
                print(f"    [debug] Falha ao salvar debug do OCR: {erro_dbg}")

        return texto
    except Exception as erro:
        print(f"[AVISO] Falha no OCR: {erro}")
        return None


def classificar_login(texto):
    """
    Classifica o resultado do login a partir do texto lido na tela.

    Detecção POSITIVA: só retorna sucesso se houver um sinal de que o usuário
    está logado (TEXTOS_LOGIN_OK). Caso contrário, retorna uma falha — evitando
    marcar como sucesso um login que na verdade deu errado.

    Retorna um dos status:
        "Login realizado com sucesso"
        "Senha incorreta"
        "Usuário inválido"
        "Falha no login"            (falhou, sem motivo específico identificado)
        "Não foi possível verificar" (OCR indisponível — não dá para afirmar)
    """
    # Sem OCR não é possível afirmar nada com segurança.
    if texto is None:
        return "Não foi possível verificar"

    # 1) Erros específicos têm prioridade (detalham o motivo).
    if any(t in texto for t in TEXTOS_SENHA_INCORRETA):
        return "Senha incorreta"
    if any(t in texto for t in TEXTOS_USUARIO_INVALIDO):
        return "Usuário inválido"

    # 2) Sinal positivo de que está logado → sucesso.
    if any(t in texto for t in TEXTOS_LOGIN_OK):
        return "Login realizado com sucesso"

    # 3) Nenhum sinal de login → considera falha (não assume sucesso).
    return "Falha no login"


def realizar_login(usuario):
    """
    ETAPA 3 — Executa o login de um usuário na tela.

    Passos: clica no campo de login → digita → clica em senha → digita →
    pressiona ENTER → aguarda confirmação.

    Retorna o STATUS resultante (ver classificar_login):
        "Login realizado com sucesso" / "Senha incorreta" /
        "Usuário inválido" / "Falha no login" / "Não foi possível verificar"
    """
    login = usuario["login"]
    senha = usuario["senha"]

    # 1-4) Campo de login e digitação.
    clicar(LOGIN_X, LOGIN_Y)
    limpar_campo()
    digitar_texto(login)

    # 5-6) Campo de senha e digitação.
    clicar(SENHA_X, SENHA_Y)
    limpar_campo()
    digitar_texto(senha)

    # 7) Confirma o login pressionando ENTER (em vez de clicar num botão).
    #    Com o cursor ainda no campo de senha, o ENTER envia o formulário.
    pyautogui.press("enter")
    time.sleep(TEMPO_ENTRE_CLIQUES)

    # 8) Aguarda para confirmar se o login foi realizado.
    time.sleep(TEMPO_APOS_LOGIN)

    # 9) Lê a tela por OCR e classifica o resultado de forma POSITIVA:
    #    só é sucesso se houver sinal de que está logado.
    texto = ler_tela(rotulo_debug=f"login_{login}")
    return classificar_login(texto)


def realizar_logout():
    """
    ETAPA 3 (continuação) — Efetua o logout e retorna à tela de login.

    Dá um único clique direto no botão "Sair" (BOTAO_SAIR_X/Y). Após o clique,
    o site retorna sozinho para a tela inicial de login, pronta para o próximo
    usuário. Eventual erro é apenas registrado (não interrompe o loop).
    """
    try:
        # 9) Clica direto no botão "Sair".
        clicar(BOTAO_SAIR_X, BOTAO_SAIR_Y)
    except Exception as erro:
        print(f"[AVISO] Falha ao tentar logout: {erro}")

    # 10) Aguarda o retorno para a tela inicial de login.
    time.sleep(TEMPO_APOS_LOGOUT)


def voltar_tela_inicial():
    """
    Garante que o navegador esteja de volta na tela inicial de login.

    Útil quando o login falha: recarrega a URL para um estado limpo antes
    do próximo usuário.
    """
    # Foca a barra de endereços (Ctrl+L), digita a URL e pressiona Enter.
    pyautogui.hotkey("ctrl", "l")
    time.sleep(0.3)
    digitar_texto(URL_SITE, pausa=0.3)
    pyautogui.press("enter")
    time.sleep(TEMPO_CARREGAMENTO_SITE)


# ===========================================================================
# RELATÓRIO
# ===========================================================================

def salvar_relatorio(resultados):
    """
    Gera o relatório final em Excel com o resultado de cada usuário.

    Colunas: Número, Nome, Login, Status.
    Salva em PASTA_RELATORIO\\resultado_validacao.xlsx.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Resultado"

    # Cabeçalho.
    sheet.append(["Número", "Nome", "Login", "Status"])

    # Uma linha por usuário processado.
    for item in resultados:
        sheet.append([
            item["numero"],
            item["nome"],
            item["login"],
            item["status"],
        ])

    os.makedirs(PASTA_RELATORIO, exist_ok=True)
    caminho_saida = os.path.join(PASTA_RELATORIO, NOME_RELATORIO)
    workbook.save(caminho_saida)
    print(f"[INFO] Relatório salvo em: {caminho_saida}")
    return caminho_saida


# ===========================================================================
# ORQUESTRAÇÃO PRINCIPAL
# ===========================================================================

def executar_automacao(limite=LIMITE_USUARIOS):
    """
    Função principal que orquestra todo o fluxo da automação.

    `limite`: quantos usuários processar (0 = todos). Útil para testes.

    Trata erros de forma resiliente: a falha em um usuário NÃO interrompe o
    processamento dos demais. Ao final, sempre tenta salvar o relatório.
    """
    print("=" * 60)
    print(" AUTOMAÇÃO DE VALIDAÇÃO DE USUÁRIOS")
    print(" (Mova o mouse para o canto superior esquerdo para abortar)")
    print("=" * 60)

    resultados = []

    # ETAPA 1 — Planilha.
    caminho_planilha = abrir_planilha()
    usuarios = ler_usuarios(caminho_planilha)

    if not usuarios:
        print("[AVISO] Nenhum usuário válido encontrado. Encerrando.")
        return

    # Aplica o limite de teste, se definido (>0): processa só os primeiros N.
    if limite and limite > 0:
        usuarios = usuarios[:limite]
        print(f"[MODO TESTE] Processando apenas {len(usuarios)} usuário(s).")

    # ETAPA 2 — Navegador e site.
    abrir_chrome()
    acessar_site()

    # ETAPA 3 — Loop por todos os usuários.
    for indice, usuario in enumerate(usuarios, start=1):
        print(f"\n[{indice}/{len(usuarios)}] Processando login: "
              f"{usuario['login']}")

        # Status inicial padrão para casos não previstos.
        status = "Erro inesperado"

        try:
            # Tenta realizar o login e captura o status retornado.
            status = realizar_login(usuario)
            print(f"    → Status: {status}")

            # Se entrou com sucesso, faz logout; caso contrário, volta à tela.
            if status == "Login realizado com sucesso":
                realizar_logout()
            else:
                # Login falhou (senha/usuário): garante tela inicial limpa.
                voltar_tela_inicial()

        except pyautogui.FailSafeException:
            # Usuário acionou o fail-safe — abortamos toda a automação.
            print("[ABORT] Fail-safe acionado. Interrompendo automação.")
            resultados.append({
                "numero": usuario["numero"],
                "nome": usuario["nome"],
                "login": usuario["login"],
                "status": "Interrompido pelo usuário",
            })
            break

        except Exception as erro:
            # Qualquer outra exceção: registra e CONTINUA para o próximo.
            print(f"[ERRO] Exceção inesperada: {erro}")
            status = "Erro inesperado"
            # Tenta recuperar o estado voltando à tela inicial.
            try:
                voltar_tela_inicial()
            except Exception:
                pass

        # Registra o resultado deste usuário no relatório.
        resultados.append({
            "numero": usuario["numero"],
            "nome": usuario["nome"],
            "login": usuario["login"],
            "status": status,
        })

    # RELATÓRIO — sempre gerado ao final.
    salvar_relatorio(resultados)
    print("\n[INFO] Automação concluída.")


# ===========================================================================
# PONTO DE ENTRADA
# ===========================================================================

if __name__ == "__main__":
    import sys

    # Limite opcional de usuários pela linha de comando (ex.: python main.py 2).
    # Se não for informado, usa LIMITE_USUARIOS (0 = todos).
    limite = LIMITE_USUARIOS
    if len(sys.argv) > 1:
        try:
            limite = int(sys.argv[1])
        except ValueError:
            print(f"[AVISO] Argumento inválido '{sys.argv[1]}'. Ignorando.")

    try:
        executar_automacao(limite=limite)
    except FileNotFoundError as erro:
        # Erro crítico (ex.: planilha não encontrada) — não há o que processar.
        print(f"[ERRO CRÍTICO] {erro}")
    except pyautogui.FailSafeException:
        print("[ABORT] Fail-safe acionado antes de iniciar o processamento.")
